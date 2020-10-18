#!/usr/bin/python
# -*- coding: UTF-8 -*-
# @Time : 2020-05-25 14:48 

# @Author : liumengjia

# @File : test_excel2.py 

# @Software: PyCharm

from openpyxl import load_workbook

expect_data = []
actual_data = []

class TestExcel():
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]

    def test_get_expc_data(self):
        for i in range(2,self.sheet.max_row+1):
            test_data = self.sheet.cell(i,7).value
            expect_data.append(test_data)
        return expect_data

    def test_get_actual_data(self):
        for j in range(2,self.sheet.max_row+1):
            sub_data = self.sheet.cell(j,9).value
            actual_data.append(sub_data)
        return actual_data

TestExcel('../data/test_clue.xlsx','assignClue').test_get_expc_data()
TestExcel('../data/test_clue.xlsx','assignClue').test_get_actual_data()

def test_response():
    # 预期结果列表，和实际结果列表断言
    assert expect_data == actual_data

e = eval(expect_data[4])
print(e)
print(type(e))

f = eval(actual_data[4])
print(f)
print(type(f))

def test_result():
    assert e == f
