#!/usr/bin/python
# -*- coding: UTF-8 -*-
# @Time : 2020-02-25 23:20 

# @Author : liumengjia

# @File : test_excel.py 

# @Software: PyCharm

from openpyxl import load_workbook

class TestExcel():
    # 使用__init__方法实现：只要实例化类时候就会调用__init__方法
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]
        self.ws = self.wb.active
        self.maxRow = self.ws.max_row
        self.maxCol = self.ws.max_column

    def test_get_data(self):
        data=[]
        '''
        根据传入的坐标来获取值
        :param i:
        :param j:
        :return:
        '''
        for i in self.sheet["A"]:
            data.append(i.value)
        return data

    def test_write_data(self,row,col,value):
        self.ws.cell(row,col,value)
        self.wb.save(self.file_name)

if __name__=='__main__':
    print(TestExcel('../data/test_addNewClue_data.xlsx','clue').test_get_data())
