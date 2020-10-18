#!/usr/bin/python
# -*- coding: UTF-8 -*-
# @Time : 2020-02-25 23:20 

# @Author : liumengjia

# @File : test_excel.py 

# @Software: PyCharm

from openpyxl import load_workbook

class TestExcel():
    # 使用__init__方法实现：只要实例化类时候就会调用__init__方法
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]
        self.ws = self.wb.active  # 激活sheet
        self.maxRow = self.ws.max_row
        self.maxCol = self.ws.max_column

    def test_get_title(self):
        '''
            获取表格第一行标题行
        '''
        title = []  # 存储标题行
        for i in range(1,self.sheet.max_column+1):   # 遍历每一列
            title.append(self.sheet.cell(1,i).value)

        return title

    def test_get_data(self):
        '''
            读取Excel表格数据
        '''
        title = self.test_get_title() #返回标题行
        test_data = []

        # 外层循环遍历每行，内层循环遍历每列
        for i in range(2,self.sheet.max_row+1):  #外层循环遍历每一行，从第2行到最大行
            sub_data = {}
            for j in range(1,self.sheet.max_column+1):  #内层循环遍历每一列，从第1列到最大列
                sub_data[title[j-1]] = self.sheet.cell(i,j).value  # sub_data[key]=value ,title从0开始
            test_data.append(sub_data)

        return test_data

    def test_write_data(self,row,col,value):
        self.ws.cell(row,col,value)

    def test_save_data(self):
        self.wb.save(self.file_name)

if __name__=='__main__':
    # 使用__init__方法实现：只要实例化类时候就会调用__init__方法

    '''写入数据，如(2,3,"test"),第二行第三列写入数据"test" '''
    TestExcel('../data/user.xlsx','accountPwdLogin').test_write_data(2,3,'test')

    print(TestExcel('../data/user.xlsx','accountPwdLogin').test_get_data())